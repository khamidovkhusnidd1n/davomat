import openpyxl
import os

file_path = r"c:\Users\Salohiddin Markaz\Downloads\Telegram Desktop\Navbatchilik iyul avgust sentabr.xlsx"
if os.path.exists(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        print("Sheet:", sheet_name)
        sheet = wb[sheet_name]
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            row_str = " | ".join([str(val) for val in row if val is not None])
            if row_str.strip():
                print(f"Row {r_idx+1}: {row_str[:200]}")
else:
    print("File not found.")
